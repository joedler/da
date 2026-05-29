import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path("outputs/行程表匯入草稿.xlsx")
ITINERARY_CSV = Path("outputs/itinerary.csv")


ITINERARY = [
    ["Day1", "2026/06/08", "一", "08:00-08:20", "學校集合、車輛安全檢查", "行、駕正本、逃生門演練", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "08:20-08:30", "師生到齊、乘車出發", "依車次上車，到齊出發", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "08:30-11:30", "國道風光", "領隊車上自我介紹，中途停留國道休息站", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "11:30-12:10", "午餐的約會", "大溪南海餐廳：七菜一湯十三瓶飲料", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "13:40-15:20", "十分風景區", "十分瀑布、車站老街，參觀100分鐘", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "16:10-18:10", "九份風景區", "芋圓、草仔粿、昇平戲院，參觀120分鐘", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "19:00-19:20", "飯店 Check in", "", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day1", "2026/06/08", "一", "19:30-21:30", "基隆廟口夜市", "不用車；食字路口大進擊，同學晚餐自理", "基隆長榮桂冠", "基隆市中正區中正路62-1號", "02-2427-9988"],
    ["Day2", "2026/06/09", "二", "06:00-06:30", "Morning Call", "早安", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "06:40-07:30", "早餐的約會", "飯店內享用自助早餐", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "07:40-07:50", "上車出發", "師生上完化妝室上車，清點人數到齊出發", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "09:00-11:00", "國立故宮博物院", "學校上網預約參觀時間，參觀2小時", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "11:10-13:00", "美麗華百樂園", "午餐發放代金 $200，於美食街自理，參觀110分鐘", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "13:30-15:00", "國立臺灣科學教育館", "三至六樓展示廳，參觀90分鐘", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "15:35-15:55", "飯店 Check in", "", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "16:00-21:20", "雙北城市探索捷運半日遊", "同學晚餐自理", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day2", "2026/06/09", "二", "21:20-21:30", "回到飯店", "領隊清點人數並發放房卡", "西門德立莊", "臺北市中正區秀山街4號", "02-2375-7777"],
    ["Day3", "2026/06/10", "三", "06:00-06:30", "Morning Call", "早安", "溫暖的家", "", ""],
    ["Day3", "2026/06/10", "三", "06:40-07:20", "早餐的約會", "飯店內享用自助早餐", "溫暖的家", "", ""],
    ["Day3", "2026/06/10", "三", "07:20-07:30", "人員及行李上車", "師生到齊出發", "溫暖的家", "", ""],
    ["Day3", "2026/06/10", "三", "09:40-13:40", "麗寶探索樂園", "參觀4小時；午餐發放代金 $200，於園區自理", "溫暖的家", "", ""],
    ["Day3", "2026/06/10", "三", "15:50-16:00", "抵達學校", "結束此次校外旅遊活動", "溫暖的家", "", ""],
]


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for col in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row, column=col).value or "")
            for row in range(1, ws.max_row + 1)
        ]
        ws.column_dimensions[get_column_letter(col)].width = min(max(map(len, values)) + 2, 50)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "itinerary"
    ws.append(["day", "date", "weekday", "time", "title", "note", "hotel", "hotel_address", "hotel_phone"])
    for row in ITINERARY:
        ws.append(row)

    summary = wb.create_sheet("手機顯示建議")
    summary.append(["區塊", "建議"])
    summary.append(["查詢頁", "新增「行程表」分頁或按鈕，依 Day1/Day2/Day3 顯示時間軸。"])
    summary.append(["顯示內容", "時間、地點/活動、重要備註、當晚住宿。"])
    summary.append(["圖片", "行程表原圖可保留在管理文件，不建議當作學生手機主要閱讀介面。"])
    summary.append(["提醒", "之後可加集合提醒，但第一版先做靜態查詢。"])

    for sheet in wb.worksheets:
        style_sheet(sheet)

    wb.save(OUT)
    with ITINERARY_CSV.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["day", "date", "weekday", "time", "title", "note", "hotel", "hotel_address", "hotel_phone"])
        writer.writerows(ITINERARY)

    print(OUT.resolve())
    print(ITINERARY_CSV.resolve())


if __name__ == "__main__":
    main()
