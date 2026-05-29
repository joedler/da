from collections import Counter, defaultdict
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC = Path("outputs/system_data.xlsx")
OUT = Path("outputs/房號匯入草稿.xlsx")
ROOM_ASSIGNMENTS_CSV = Path("outputs/room_assignments.csv")


def group_no(group: str) -> int:
    digits = "".join(ch for ch in str(group) if ch.isdigit())
    return int(digits or 0)


def main_class(rows: list[list[object]]) -> str:
    counts = Counter(str(row[3]) for row in rows)
    return counts.most_common(1)[0][0] if counts else ""


def members(rows: list[list[object]]) -> str:
    return "、".join(str(row[2]) for row in rows)


def roles(rows: list[list[object]]) -> str:
    return "、".join(f"{row[2]}({row[3]})" for row in rows)


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for col in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row, column=col).value or "")
            for row in range(1, min(ws.max_row, 60) + 1)
        ]
        ws.column_dimensions[get_column_letter(col)].width = min(max(map(len, values)) + 2, 48)


def main() -> None:
    source_wb = load_workbook(SRC, data_only=True)
    people_ws = source_wb["people"]
    rows = [list(row) for row in people_ws.iter_rows(min_row=2, values_only=True) if row[2]]

    rooms: dict[str, list[list[object]]] = defaultdict(list)
    for row in rows:
        if row[11]:
            rooms[str(row[11])].append(row)

    # Confirmed room pools from the PDFs and user decisions.
    d2_male_3 = ["0628", "0630", "0646", "0648", "0663", "0669", "0671", "0730", "0769"]
    d2_male_4 = [
        "0601", "0602", "0603", "0606", "0610", "0616", "0618", "0622", "0626", "0632",
        "0636", "0638", "0640", "0642", "0702", "0706", "0708", "0710", "0716", "0718",
        "0722", "0726", "0732", "0736", "0738", "0740", "0742", "0746", "0748",
    ]
    d2_female_3 = ["0508", "0542"]
    d2_female_4 = ["0502", "0506", "0530", "0536", "0538"]
    d2_teacher_regular = ["0101", "0105", "0107", "0109", "0110", "0115", "0119"]
    d2_teacher_special = {"師1": "0775", "師11": "0515", "師9": "0615", "師10": "0715"}

    d1_female_3 = ["1301", "1310"]
    d1_female_4 = ["1302", "1303", "1304", "1307", "1308"]
    d1_teacher = ["0805", "0806", "0807", "0902", "0904", "0905", "0906", "1005", "1006", "1106"]
    d1_teacher_map = {
        "師1": "長榮不住宿",
        "師2": "0806",
        "師6": "0805",
        "師3": "0807",
        "師4": "0902",
        "師5": "0904",
        "師7": "0905",
        "師8": "0906",
        "師9": "1005",
        "師10": "1006",
        "師11": "1106",
    }
    d1_male_3 = ["1101", "0911", "0915", "1011", "1015", "1115", "1211", "1411", "1415"]
    d1_male_4 = [
        "1002", "1003", "1004", "1007", "1008", "1009", "1102", "1103", "1104",
        "1107", "1108", "1109", "1207", "1208",
        "0912", "0913", "0914", "1012", "1013", "1014", "1112", "1113", "1114",
        "1212", "1213", "1412", "1414",
    ]

    d2_map: dict[str, str] = {}
    for prefix, room3, room4 in [("男", d2_male_3, d2_male_4), ("女", d2_female_3, d2_female_4)]:
        i3 = i4 = 0
        for group in sorted([g for g in rooms if g.startswith(prefix)], key=group_no):
            size = len(rooms[group])
            if size == 3:
                d2_map[group] = room3[i3] if i3 < len(room3) else "待補"
                i3 += 1
            elif size == 4:
                d2_map[group] = room4[i4] if i4 < len(room4) else "待補"
                i4 += 1
            else:
                d2_map[group] = "人數異常"

    regular_teacher_rooms = iter(d2_teacher_regular)
    for group in sorted([g for g in rooms if g.startswith("師")], key=group_no):
        d2_map[group] = d2_teacher_special.get(group) or next(regular_teacher_rooms, "待補")

    d1_map: dict[str, str] = {}
    i3 = i4 = 0
    for group in sorted([g for g in rooms if g.startswith("女")], key=group_no):
        size = len(rooms[group])
        if size == 3:
            d1_map[group] = d1_female_3[i3] if i3 < len(d1_female_3) else "待補"
            i3 += 1
        elif size == 4:
            d1_map[group] = d1_female_4[i4] if i4 < len(d1_female_4) else "待補"
            i4 += 1

    for group in sorted([g for g in rooms if g.startswith("師")], key=group_no):
        d1_map[group] = d1_teacher_map.get(group, "待補")

    i3 = i4 = 0
    for group in sorted([g for g in rooms if g.startswith("男")], key=group_no):
        size = len(rooms[group])
        if size == 3:
            d1_map[group] = d1_male_3[i3] if i3 < len(d1_male_3) else "待補"
            i3 += 1
        elif size == 4:
            d1_map[group] = d1_male_4[i4] if i4 < len(d1_male_4) else "待補"
            i4 += 1

    wb = Workbook()
    draft_ws = wb.active
    draft_ws.title = "房號匯入草稿"
    draft_ws.append([
        "房間編組", "類別", "人數", "主要班級或職稱", "同房人員",
        "第一天飯店", "第一天房號", "第二天飯店", "第二天房號", "備註",
    ])

    def display_sort_key(group: str) -> tuple[int, int]:
        if group.startswith("師"):
            return (0, group_no(group))
        if group.startswith("男"):
            return (1, group_no(group))
        return (2, group_no(group))

    for group in sorted(rooms, key=display_sort_key):
        group_rows = rooms[group]
        note = ""
        if group == "師1":
            note = "校長長榮不住宿；德立莊顯示房號。"
        if group in {"師2", "師6"}:
            note = "長榮維持單人一間。"
        if group.startswith("男"):
            note = "兩晚皆照原房間編組。"

        draft_ws.append([
            group,
            group_rows[0][1],
            len(group_rows),
            main_class(group_rows),
            members(group_rows),
            "長榮桂冠酒店",
            d1_map.get(group, ""),
            "德立莊",
            d2_map.get(group, ""),
            note,
        ])

    decision_ws = wb.create_sheet("長榮男生待決策")
    decision_ws.append(["項目", "數量", "說明"])
    male_groups = [g for g in rooms if g.startswith("男")]
    decision_ws.append(["目前男生房間編組", len(male_groups), "現有系統分房編組"])
    decision_ws.append(["目前男生三人組", sum(1 for g in male_groups if len(rooms[g]) == 3), "若維持同房，需要 9 間三人房或可住 3 人的房"])
    decision_ws.append(["目前男生四人組", sum(1 for g in male_groups if len(rooms[g]) == 4), "若維持同房，需要 29 間四人房"])
    decision_ws.append(["長榮男生三人房", len(d1_male_3), "依 2026/05/29 確認：9 間"])
    decision_ws.append(["長榮男生四人房", len(d1_male_4), "依 2026/05/29 確認：29 間"])
    decision_ws.append(["結論", "", "長榮男生已調整為 29 間四人房 + 9 間三人房，兩晚皆照原房間編組。"])
    decision_ws.append([])
    decision_ws.append(["房間編組", "人數", "主要班級", "成員"])
    for group in sorted(male_groups, key=group_no):
        group_rows = rooms[group]
        decision_ws.append([group, len(group_rows), main_class(group_rows), roles(group_rows)])

    pool_ws = wb.create_sheet("房號池")
    pool_ws.append(["飯店", "對象", "房型/人數", "房號"])
    for room in d1_female_3:
        pool_ws.append(["長榮桂冠酒店", "女生", "3人房", room])
    for room in d1_female_4:
        pool_ws.append(["長榮桂冠酒店", "女生", "4人房", room])
    for room in d1_teacher:
        pool_ws.append(["長榮桂冠酒店", "老師", "2人房", room])
    for room in d1_male_3:
        pool_ws.append(["長榮桂冠酒店", "男生", "3人房", room])
    for room in d1_male_4:
        pool_ws.append(["長榮桂冠酒店", "男生", "4人房", room])
    for room in d2_female_3:
        pool_ws.append(["德立莊", "女生", "3人房", room])
    for room in d2_female_4:
        pool_ws.append(["德立莊", "女生", "4人房", room])
    for room in d2_male_3:
        pool_ws.append(["德立莊", "男生", "3人房", room])
    for room in d2_male_4:
        pool_ws.append(["德立莊", "男生", "4人房", room])
    for room in ["0775"] + d2_teacher_regular + ["0515", "0615", "0715"]:
        pool_ws.append(["德立莊", "老師/校長", "老師房", room])

    for sheet in wb.worksheets:
        style_sheet(sheet)

    wb.save(OUT)
    with ROOM_ASSIGNMENTS_CSV.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["房間編組", "第一天飯店", "第一天房號", "第二天飯店", "第二天房號", "備註"])
        for row in draft_ws.iter_rows(min_row=2, values_only=True):
            writer.writerow([row[0], row[5], row[6], row[7], row[8], row[9]])

    print(OUT.resolve())
    print(ROOM_ASSIGNMENTS_CSV.resolve())


if __name__ == "__main__":
    main()
