import csv
import json
from pathlib import Path

from openpyxl import load_workbook


OUT = Path("outputs/contacts.csv")
SYSTEM_DATA = Path("outputs/system_data.xlsx")
PRIVATE_INFO = Path("outputs/private_info.json")

# Default mock data for public repo / fallback
DEFAULT_BUS_LEADERS = [
    ["旅行社領隊", "領隊A", "A車", "0900000001", "", ""],
    ["旅行社領隊", "領隊B", "B車", "0900000002", "", ""],
    ["旅行社領隊", "領隊C", "C車", "0900000003", "", ""],
    ["旅行社領隊", "領隊D", "D車", "0900000004", "", ""],
    ["旅行社領隊", "領隊E", "E車", "0900000005", "", ""],
    ["旅行社領隊", "領隊F", "F車", "0900000006", "", ""],
]
DEFAULT_AGENCY_CONTACT = ["旅行社窗口", "聯絡窗口", "", "0900000000", "", "旅行社窗口"]


def load_private_info() -> tuple[list[list[str]], list[str]]:
    if PRIVATE_INFO.exists():
        try:
            with PRIVATE_INFO.open("r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("bus_leaders", DEFAULT_BUS_LEADERS), data.get("agency_contact", DEFAULT_AGENCY_CONTACT)
        except Exception:
            pass
    return DEFAULT_BUS_LEADERS, DEFAULT_AGENCY_CONTACT


def school_contacts() -> list[list[str]]:
    if not SYSTEM_DATA.exists():
        return []

    workbook = load_workbook(SYSTEM_DATA, data_only=True)
    sheet = workbook["people"]
    contacts: list[list[str]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        person_type = str(row[1] or "")
        name = str(row[2] or "")
        role = str(row[3] or "")
        bus = str(row[8] or "")
        if person_type != "老師" or not name:
            continue

        if "導師" in role:
            contacts.append(["班導師", name, bus, "", "", role])
        elif role == "護理師":
            contacts.append(["護理師", name, "", "", "", "全團護理支援"])
        else:
            contacts.append(["學校窗口", name, bus, "", "", role])

    return contacts


def main() -> None:
    bus_leaders, agency_contact = load_private_info()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["類型", "姓名", "服務車次", "電話", "LINE或其他聯絡方式", "備註"])
        writer.writerows(bus_leaders)
        writer.writerows(school_contacts())
        writer.writerow(agency_contact)
    print(OUT.resolve())


if __name__ == "__main__":
    main()

